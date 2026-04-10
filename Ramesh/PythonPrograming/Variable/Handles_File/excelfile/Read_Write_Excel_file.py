# install library with command on terminal
# -> pip install openpyxl

import openpyxl

def read_excel_file(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


# read_excel_file(file_path="user_data.xlsx", sheet_name="Sheet1", cell_name="A1")
excel_file_path = r"C:\GitCode\GTM_PS_BATCH16\Ramesh\PythonPrograming\Variable\Handles_File\user_data.xlsx.xlsx"

read_excel_file(file_path=excel_file_path, sheet_name="Sheet1", cell_name="A1")

print("--"*20)

# for i in range(1, 7):
#    read_excel_file(file_path=excel_file_path, sheet_name="Sheet1", cell_name=f"A{i}")


def read_excel_file(file_path, sheet_name):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    # get maximum number of rows in sheet
    max_row_value = sheet.max_row
    # get maximum number of rows in sheet
    max_colum_value = sheet.max_column
    # i consider as row number

    for i in range(1, max_row_value):

        # j consider as a colum number

        for j in range(1, max_colum_value):
            print(sheet.cell(row=i, column=j).value, end= " ")
        print()
        
excel_file_path = r"C:\GitCode\GTM_PS_BATCH16\Ramesh\PythonPrograming\Variable\Handles_File\user_data.xlsx.xlsx"
read_excel_file(file_path=excel_file_path, sheet_name="Sheet2")

print("___"*10)
def write_excel_file(file_path, sheet_name, cell_name, cell_value):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = cell_value
    wb.save(file_path)


# write_excel_file(file_path=excel_file_path, sheet_name="Sheet3",cell_name="A1", cell_value="INDIA")