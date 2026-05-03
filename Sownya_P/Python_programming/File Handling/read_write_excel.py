# install openpyxl library to read/write excel files
# pip install openpyxl


import openpyxl

def read_excel(file_path, sheet_name, cell_name):
    # loading the workbook and selecting the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Read specific cell value
    cell_value = sheet[cell_name].value
    print("Value in A1:", cell_value)

     # Read entire column
    print("Values in column A:")
    for cell in sheet['A']:
        print(cell.value)
        # Read specific row
      
    # Read entire row
    print("Values in row 1:")
    for cell in sheet[1]:
        print(cell.value)

read_excel(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet1", cell_name="A1")    
   

def write_excel(file_path, sheet_name, cell_name, value):
    # loading the workbook and selecting the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Write value to specific cell
    sheet[cell_name] = value

    # Save the workbook
    workbook.save(file_path)
    print(f"Value '{value}' has been written to {cell_name} in {sheet_name} of {file_path}")

write_excel(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet1", cell_name="D1", value="python learning")
##write_excel(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet2", cell_name="A1", value="othersheet")

data = ["Name", "Age", "City", "Email", "Phone"]

for index,data in enumerate(data, start=1):
    write_excel(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet2", cell_name=f"E{index+1}", value=data)


credentials = [["username", "password"], ["sownya", "password123"], ["Pavan", "password456"], ["Chala", "password789"]]

for index, val in enumerate(credentials):
     write_excel(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet2", cell_name=f"A{index+1}",value=val[0])
     write_excel(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet2", cell_name=f"B{index+1}",value=val[1])


#iterate through rows and columns to read/write data
def read_excel_with_iteration(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(values_only=True):
        print(row)
 # New line after each row
print("*" * 20 )
read_excel_with_iteration(file_path=r"C:\AutomationProject\GTM_PS_BATCH16\Sownya_P\Python programming\File Handling\user_Data.xlsx", sheet_name="Sheet3")