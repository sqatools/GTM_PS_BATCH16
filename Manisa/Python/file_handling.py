#install library with command on terminal
#pip install openpyxl

import openpyxl #excel libraries

def read_excel_file(file_path, sheet_name, cell_name):
    wb =openpyxl.load_workbook(filename = file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)

excel_file_path = r"C:\PythonAutomation\GTM_PS_BATCH16\Manisa\excelfile.xlsx"
read_excel_file(file_path = excel_file_path, sheet_name="Sheet1", cell_name="A2")

#for i in range (1, 6):
    #read_excel_file(file_path=excel_file_path, sheet_name="Sheet1", cell_name=f"A{i}")

def read_excelfile(file_path, sheet_name):
    wb =openpyxl.load_workbook(filename = file_path)
    sheet = wb[sheet_name]
    max_row_value = sheet.max_row
    print("max row value:", max_row_value)
    max_column_value = sheet.max_column
    print("max column value:", max_column_value)
    for i in range(1, max_row_value+1):
        for j in range(1, max_column_value+1):
            print(sheet.cell(row=i, column=j).value, end=" ")
        print()

read_excelfile(file_path=excel_file_path, sheet_name="Sheet1")

'''def write_excel_file(file_path, sheet_name, cell_name, cell_value):
    wb =openpyxl.load_workbook(filename = file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = cell_value
    wb.save(file_path)
    #print(cell.value)

write_excel_file(file_path=excel_file_path, sheet_name="Sheet2", cell_name="A1", cell_value="india")'''