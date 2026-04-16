# openpyxl is external library to interact with excel file

import openpyxl

#def read_excel_file(file_path,sheet_name,cell_name):
   # wb=openpyxl.load_workbook(filename=file_path)
  #  sheet=wb[sheet_name]
 #   cell=sheet[cell_name]
#    print(cell.value)



#read_excel_file(file_path="employee_data.xlsx",sheet_name="Sheet1",cell_name="A4")

#for i in range(1, 7):
 #   read_excel_file(file_path="employee_data.xlsx",sheet_name="Sheet1",cell_name=f"A{i}")

def read_excel_file(file_path,sheet_name):
    wb=openpyxl.load_workbook(filename=file_path)
    sheet= wb[sheet_name]
    max_row_value=sheet.max_row
    max_column_value=sheet.max_column
    for i in range(1,max_row_value):
        for j in range(1,max_column_value):
            print(sheet.cell(row=i,column=j).value, end= "")
            print()

 
 
read_excel_file(file_path="employee_data.xlsx",sheet_name="Sheet1")


 



