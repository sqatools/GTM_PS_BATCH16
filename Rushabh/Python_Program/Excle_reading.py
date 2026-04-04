import openpyxl

def read_excel_file(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)       

excel_file_path = r"C:\GitCode\GTM_PS_BATCH16\Rushabh\Python_Program\New_user_data.xlsx"
read_excel_file(file_path= excel_file_path, sheet_name="Sheet1", cell_name="A1")

def read_excel_file(file_path, sheet_name):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    # get max number of updated rows in sheet
    max_row_value = sheet.max_row
    # get max number of updated colms in sheet
    max_colum_value = sheet.max_column
    # i consider as row number
    for i in range(1, max_row_value):
        # j consider as colum number
        for j in range(1, max_colum_value):
            print(sheet.cell(row=i, column=j).value, end=" ")
        print()                     

#read_excel_file(file_path=excel_file_path, sheet_name="Sheet2")    

def write_excel_file(file_path, sheet_name, cell_name, cell_value):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = cell_value
    wb.save(file_path)                  

#write_excel_file(file_path=excel_file_path, sheet_name="Sheet3", cell_name="A1", cell_value="India")    