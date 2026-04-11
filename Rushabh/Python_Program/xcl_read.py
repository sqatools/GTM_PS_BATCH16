import openpyxl

def read_xcl_file(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)
  #  wb = openpyxl.load_workbook(r"C:\GitCode\GTM_PS_BATCH16\Rushabh\Python_Program\Demo_Xcl.xlsx")
  #  print(wb.sheetnames)

excel_file_path = r"C:\GitCode\GTM_PS_BATCH16\Rushabh\Python_Program\Demo_Xcl.xlsx"
read_xcl_file(file_path=excel_file_path, sheet_name="Sheet1", cell_name="A1")


def read_xcl_file(file_path, sheet_name):
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    # get max number of updated rows in sheet

    max_row_value = sheet.max_row
    # get max number of updated colms in sheet

    max_colum_value = sheet.max_column
    # i consider as row number
    
    for i in range(1, max_row_value):
        # j consider as colum number
        for j in range(1, max_colum_value):
            print(sheet.cell(row=i, column=j).value, end=" ")
        print()             
#excel_file_path = r"C:\GitCode\GTM_PS_BATCH16\Rushabh\Python_Program\Demo_Xcl.xlsx"
read_xcl_file(file_path=excel_file_path, sheet_name="Sheet2")